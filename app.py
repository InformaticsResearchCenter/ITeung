# -*- coding: utf-8 -*-
"""
Created on Sat Mar 14 09:27:37 2020

@author: rolly
"""

from lib import iteung, wa
from flask import Flask, request, render_template, make_response, jsonify, send_from_directory
from twilio.twiml.messaging_response import MessagingResponse
from lib import log
from module import kelas
from Crypto.Cipher import AES
from base64 import b64decode
from datetime import datetime
from module import surat_va
from openpyxl import load_workbook

from flask_restful import Resource, Api, abort

import config, pymysql, flask_chatbot

app = Flask(__name__, static_url_path='')
apirest=Api(app=app)

def dbConnectPMB():
    db=pymysql.connect(config.db_host_pmb, config.db_username_pmb, config.db_password_pmb, config.db_name_pmb)
    return db

def insertnewKHS(npm, tahunid, prodiid, tipesemester, biaya):
    db=kelas.dbConnectSiap()
    sql=f"INSERT INTO simak_trn_khs (KHSID,TahunID,ProdiID,KodeID,ProgramID,MhswID,StatusMhswID,sesi,MaxSKS,Cetak,Biaya) VALUES (DEFAULT,'{tahunid}','{prodiid} ','YPBPI','REG','{npm}','A','{tipesemester}','24','Y', {biaya});"
    with db:
        cur=db.cursor()
        cur.execute(sql)

def updateBiayaKHS(npm, tahunid, biaya):
    db=kelas.dbConnectSiap()
    sql=f"UPDATE simak_trn_khs SET Biaya={biaya} WHERE MhswID={npm} and TahunID={tahunid}"
    with db:
        cur=db.cursor()
        cur.execute(sql)

def cekSudahAdaKHS(npm, tahunid, statusmahasiswa):
    db=kelas.dbConnectSiap()
    sql=f"select * from simak_trn_khs where MhswID={npm} and TahunID={tahunid} and StatusMhswID='{statusmahasiswa}'"
    with db:
        cur=db.cursor()
        cur.execute(sql)
        row=cur.fetchone()
        if row is not None:
            return True
        else:
            return False

def cekSesiSemester(tipesemester, npm):
    tahunangkatan=int(kelas.getTahunAngkatanWithStudentID(npm))
    tahunsekarang=int(datetime.now().strftime('%Y'))
    selisihtahun=tahunsekarang-tahunangkatan
    if tipesemester == '1':
        selisihtahun+=0.5
        sesisemester=selisihtahun/0.5
    else:
        selisihtahun+=1
        sesisemester=selisihtahun/0.5
    return str(sesisemester)

def floatToRupiah(uang):
    uang=float(uang)
    str_uang=str(uang)
    uang_split=str_uang.split('.')[0]
    uang_reverse=uang_split[::-1]
    uang_temp=''
    for i, j in enumerate(uang_reverse):
        if i%3==0:
            uang_temp+=f'.{j}'
        else:
            uang_temp+=f'{j}'
    uang_reverse_jadi=uang_temp[1:]
    return f'Rp {uang_reverse_jadi[::-1]},{str_uang.split(".")[1]}'

def decryptToken(key, iv, passcode):
    try:
        key_tobytes=bytes(key, 'utf-8')
        iv_tobytes=bytes(iv, 'utf-8')
        crypt_object=AES.new(key=key_tobytes,mode=AES.MODE_CBC,IV=iv_tobytes)
        passcode=b64decode(passcode)
        ciphertext=passcode
        resultpasscode, status = crypt_object.decrypt(b64decode(ciphertext)), True
    except Exception as e:
        resultpasscode, status = f'', False
    return str(resultpasscode, 'utf-8').replace('\x0e', ''), status


def cekNpmInTrxID(trxid):
    data=trxid.split('-')
    for i in data:
        if kelas.getDataMahasiswa(i):
            return i


def cekTipeSemester(trxid):
    for i in trxid.split('-'):
        try:
            int(i)
            isNumber=True
        except:
            isNumber=False
        if len(i) == 1 and isNumber:
            return i


def openfile():
    namafile='wekwek.xlsx'
    wb = load_workbook(namafile)
    return wb


def getDataDefault(key, ws):
    switcher={
        'd3titk22019': ws['F8'].value,
        'd3titk32018': ws['H8'].value,
        'd3mitk22019': ws['F9'].value,
        'd3mitk32018': ws['H9'].value,
        'd3aktk22019': ws['F10'].value,
        'd3aktk32018': ws['H10'].value,
        'd3mbtk22019': ws['F11'].value,
        'd3mbtk32018': ws['H11'].value,
        'd3lbtk22019': ws['F12'].value,
        'd3lbtk32018': ws['H12'].value,
        'd4titk22019': ws['F21'].value,
        'd4titk32018': ws['H21'].value,
        'd4titk42017': ws['J21'].value,
        'd4aktk22019': ws['F22'].value,
        'd4aktk32018': ws['H22'].value,
        'd4aktk42017': ws['J22'].value,
        'd4mbtk22019': ws['F23'].value,
        'd4mbtk32018': ws['H23'].value,
        'd4mbtk42017': ws['J23'].value,
        'd4lbtk22019': ws['F24'].value,
        'd4lbtk32018': ws['H24'].value,
        'd4lbtk42017': ws['J24'].value
    }
    return switcher.get(key, 'not found!!!')


def getProdiSingkatanFromProdiID(prodiid):
    db=kelas.dbConnectSiap()
    sql=f'select Singkatan from simak_mst_prodi where ProdiID={prodiid}'
    with db:
        cur=db.cursor()
        cur.execute(sql)
        row=cur.fetchone()
        return row[0]


@app.route("/", methods=['GET', 'POST'])
def home():
    email=request.cookies.get('email')
    if email:
        if request.method == 'GET':
            return render_template('chatbot.html', bot_reply='')
        if request.method == 'POST':
            message=request.form['message']
            return render_template('chatbot.html', bot_reply=flask_chatbot.cekAndSendMessage(message))
    else:
        return render_template('login.html')

@app.route("/wa", methods=['GET', 'POST'])
def sms_reply():
    msg = request.values.get('Body', None)
    wanum = request.values.get('From', None)
    num=wanum.split('+')[1]
    msgreply=iteung.get(num,msg)
    resp = MessagingResponse()
    resp.message(msgreply)
    return str(resp)

@app.route('/<name>')
def senddatajavascript(name):
    return render_template('index.html', groupname=name)

@app.route('/data/proses', methods=['POST'])
def prosesdata():
    req = request.get_json()
    number=req['number']
    url=req['url']
    groupname = url.split('/')[-1].replace('%20', ' ')
    message = 'hadir'
    alias = kelas.getNpmandNameMahasiswa(number)[1]
    isgroup='true'
    tipe='luring'
    log.save(number, message, alias, groupname, isgroup, tipe)
    res = make_response(jsonify({'message': 'JSON data received'}), 200)
    return res

@app.route('/<token>/callback/api/va', methods=['POST'])
def callback_api_va(token):
    try:
        datenow = datetime.date(datetime.now()).strftime('%d-%m-%Y')
        yearnow=datetime.date(datetime.now()).strftime('%Y')
        req=request.json
        trxid=req['trx_id']
        npm=cekNpmInTrxID(trxid)
        tipesemester=cekTipeSemester(trxid)
        tahunid=f'{yearnow}{tipesemester}'
        prodiid=f'{npm[0]}{npm[3]}'
        virtual_account=req['virtual_account']
        customer_name=req['customer_name']
        trx_amount=req['trx_amount']
        payment_amount=req['payment_amount']
        cumulative_payment_amount=req['cumulative_payment_amount']
        payment_ntb=req['payment_ntb']
        datetime_payment=req['datetime_payment']
        datetime_payment_iso8601=req['datetime_payment_iso8601']
        resultpasscode, status = decryptToken(config.key_va, config.iv_va, token)
        if 'SPP' in trxid.split('-'):
            if status:
                passcodetrxid=resultpasscode.split(';')[0].replace('\n', '').replace(' ', '')
                passcodevirtualaccount=resultpasscode.split(';')[1].replace('\n', '').replace(' ', '')
                passcodedatetime=resultpasscode.split(';')[2].replace('\n', '').replace(' ', '')
                if passcodetrxid == trxid and passcodevirtualaccount == virtual_account and passcodedatetime == datenow:
                    message = f'Hai haiiiii kamu sudah transfer pembayaran semester yaaaa dengan{config.whatsapp_api_lineBreak}{config.whatsapp_api_lineBreak}*NPM: {npm}*{config.whatsapp_api_lineBreak}*Nama: {customer_name}*{config.whatsapp_api_lineBreak}*Virtual Account: {virtual_account}*{config.whatsapp_api_lineBreak}*Tanggal: {datetime_payment}*{config.whatsapp_api_lineBreak}*Jumlah Transfer: {floatToRupiah(payment_amount)}*{config.whatsapp_api_lineBreak}*Total Sudah Bayar: {floatToRupiah(cumulative_payment_amount)}*{config.whatsapp_api_lineBreak}*Total Harus Bayar: {floatToRupiah(trx_amount)}*'
                    ws = openfile().active
                    prodi_singkatan = getProdiSingkatanFromProdiID(kelas.getProdiIDwithStudentID(npm)).lower()
                    tingkat = f"tk{int(datetime.now().strftime('%Y')) - int(kelas.getTahunAngkatanWithStudentID(npm)) + 1}"
                    angkatan = kelas.getTahunAngkatanWithStudentID(npm)
                    key = f'{prodi_singkatan}{tingkat}{angkatan}'
                    default_amount_payment = getDataDefault(key, ws)
                    if int(trx_amount) > int(default_amount_payment):
                        amount_tunggakan = int(trx_amount) - int(default_amount_payment)
                        fifty_percent_default_payment = int(default_amount_payment) / 2
                        minimum_payment = int(amount_tunggakan) + int(fifty_percent_default_payment)
                    else:
                        minimum_payment = int(trx_amount) / 2
                    openfile().close()
                    if float(cumulative_payment_amount) >= float(minimum_payment):
                        if cekSudahAdaKHS(npm, tahunid, 'A'):
                            updateBiayaKHS(npm, tahunid, trx_amount-cumulative_payment_amount)
                            message += f'{config.whatsapp_api_lineBreak}{config.whatsapp_api_lineBreak}terima kasih yaaa sudah bayar semester, semangat kuliahnya kakaaaa......'
                        else:
                            message += f'{config.whatsapp_api_lineBreak}{config.whatsapp_api_lineBreak}Kamu *sudah bisa* isi KRS yaaa coba cek di *SIAP* yaaa...., #BOTNAME# ucapkan terima kasihhhh dan jangan salah saat isi KRS yaaa....'
                            message = message.replace('#BOTNAME#', config.bot_name)
                            insertnewKHS(npm, tahunid, prodiid, cekSesiSemester(tipesemester, npm), trx_amount-cumulative_payment_amount)
                        wa.setOutbox(kelas.getHandphoneMahasiswa(npm), message)
                        return make_response(jsonify(
                            {
                                "message": "success",
                                "status": "krs otomatis"
                            }
                        ), 200)
                    else:
                        message+=f'{config.whatsapp_api_lineBreak}{config.whatsapp_api_lineBreak}Yahhhh kamu *belum bisa* isi KRS nihhhh coba *buat surat* lalu *ajukan ke pihak BAUK* agar kamu bisa isi KRS..... Suratnya udah {config.bot_name} kirim ke *{kelas.getStudentEmail(npm)}*'
                        wa.setOutbox(kelas.getHandphoneMahasiswa(npm), message)
                        tes=surat_va.makePdfAndSendToEmail(npm)
                        return make_response(jsonify(
                            {
                                "message": "success",
                                "status": "kirim whatsapp bikin surat"
                            }
                        ), 200)
                else:
                    return make_response(jsonify({'message': 'bad token'}), 401)
            else:
                return make_response(jsonify({'message': 'bad token'}), 401)
        else:
            return make_response(jsonify({'message': 'not spp'}), 200)
    except Exception as e:
        return {
            "code": 404,
            "message": f"ERROR={e}"
        }

class PMDK(Resource):
    def get(self, token):
        if token == config.tokenpmb:
            db=dbConnectPMB()
            query='SELECT nama_lengkap, jenis_kelamin, nisn, ttl, telephone, ' \
                  'bbm_line, email, nama_ayah_kandung, hp_ayah_kandung, nama_ibu_kandung, ' \
                  'hp_ibu_kandung,informasi_kampus, guru_bk, hp_guru_bk, status_kelulusan, ' \
                  'daftar_ulang, date_daftar_ulang, nama_penyetor, nominal_yang_disetor, tanggal_penyetoran ' \
                  'from mahasiswa_baru where tahun = "2020/2021" and jalur = "pmdk"'
            with db:
                cur=db.cursor()
                cur.execute(query)
                data=cur.fetchall()
                if data:
                    columnname = [headers[0] for headers in cur.description]
                    json_data = []
                    for result in data:
                        json_data.append(dict(zip(columnname, result)))
                    data = {'data_pmdk': json_data}
                    response=jsonify(data)
                    response.status_code=200
                    return response
                else:
                    abort(404, message= "data kosong...")
        else:
            abort(401, message='bad token')

class Reguler(Resource):
    def get(self, token):
        if token == config.tokenpmb:
            db = dbConnectPMB()
            query = 'SELECT nama_lengkap, jenis_kelamin, nisn, ttl, telephone, bbm_line, ' \
                    'email, nama_ayah_kandung, hp_ayah_kandung, nama_ibu_kandung, hp_ibu_kandung,' \
                    'informasi_kampus, guru_bk, hp_guru_bk, status_kelulusan, daftar_ulang, ' \
                    'date_daftar_ulang, nama_penyetor, nominal_yang_disetor, tanggal_penyetoran ' \
                    'from mahasiswa_baru where tahun = "2020/2021" and jalur = "reguler"'
            with db:
                cur = db.cursor()
                cur.execute(query)
                data = cur.fetchall()
                if data:
                    columnname = [headers[0] for headers in cur.description]
                    json_data = []
                    for result in data:
                        json_data.append(dict(zip(columnname, result)))
                    data = {'data_reguler': json_data}
                    response = jsonify(data)
                    response.status_code = 200
                    return response
                else:
                    abort(404, message="data kosong...")
        else:
            abort(401, message='bad token')

class Mandiri(Resource):
    def get(self, token):
        if token == config.tokenpmb:
            db = dbConnectPMB()
            query = 'SELECT nama_lengkap, jenis_kelamin, nisn, ttl, telephone, bbm_line, ' \
                    'email, nama_ayah_kandung, hp_ayah_kandung, nama_ibu_kandung, hp_ibu_kandung,' \
                    'informasi_kampus, guru_bk, hp_guru_bk, status_kelulusan, daftar_ulang, ' \
                    'date_daftar_ulang, nama_penyetor, nominal_yang_disetor, tanggal_penyetoran ' \
                    'from mahasiswa_baru where tahun = "2020/2021" and jalur = "mandiri"'
            with db:
                cur = db.cursor()
                cur.execute(query)
                data = cur.fetchall()
                if data:
                    columnname = [headers[0] for headers in cur.description]
                    json_data = []
                    for result in data:
                        json_data.append(dict(zip(columnname, result)))
                    data = {'data_mandiri': json_data}
                    response = jsonify(data)
                    response.status_code = 200
                    return response
                else:
                    abort(404, message="data kosong...")
        else:
            abort(401, message='bad token')

class Undangan(Resource):
    def get(self, token):
        if token == config.tokenpmb:
            db = dbConnectPMB()
            query = 'SELECT nama_lengkap, jenis_kelamin, nisn, ttl, telephone, bbm_line, ' \
                    'email, nama_ayah_kandung, hp_ayah_kandung, nama_ibu_kandung, hp_ibu_kandung,' \
                    'informasi_kampus, guru_bk, hp_guru_bk, status_kelulusan, daftar_ulang, ' \
                    'date_daftar_ulang, nama_penyetor, nominal_yang_disetor, tanggal_penyetoran ' \
                    'from mahasiswa_baru where tahun = "2020/2021" and jalur = "undangan"'
            with db:
                cur = db.cursor()
                cur.execute(query)
                data = cur.fetchall()
                if data:
                    columnname = [headers[0] for headers in cur.description]
                    json_data = []
                    for result in data:
                        json_data.append(dict(zip(columnname, result)))
                    data = {'data_mandiri': json_data}
                    response = jsonify(data)
                    response.status_code = 200
                    return response
                else:
                    abort(404, message="data kosong...")
        else:
            abort(401, message='bad token')

class Beasiswa(Resource):
    def get(self, token):
        if token == config.tokenpmb:
            db = dbConnectPMB()
            query = 'SELECT nama_lengkap, jenis_kelamin, nisn, ttl, telephone, bbm_line, ' \
                    'email, nama_ayah_kandung, hp_ayah_kandung, nama_ibu_kandung, hp_ibu_kandung,' \
                    'informasi_kampus, guru_bk, hp_guru_bk, status_kelulusan, daftar_ulang, ' \
                    'date_daftar_ulang, nama_penyetor, nominal_yang_disetor, tanggal_penyetoran ' \
                    'from mahasiswa_baru where tahun = "2020/2021" and jalur = "beasiswa"'
            with db:
                cur = db.cursor()
                cur.execute(query)
                data = cur.fetchall()
                if data:
                    columnname = [headers[0] for headers in cur.description]
                    json_data = []
                    for result in data:
                        json_data.append(dict(zip(columnname, result)))
                    data = {'data_beasiswa': json_data}
                    response = jsonify(data)
                    response.status_code = 200
                    return response
                else:
                    abort(404, message="data kosong...")
        else:
            abort(401, message='bad token')

apirest.add_resource(PMDK, '/<string:token>/api/android/pmdk')
apirest.add_resource(Reguler, '/<string:token>/api/android/reguler')
apirest.add_resource(Mandiri, '/<string:token>/api/android/mandiri')
apirest.add_resource(Undangan, '/<string:token>/api/android/undangan')
apirest.add_resource(Beasiswa, '/<string:token>/api/android/beasiswa')

if __name__ == "__main__":
    app.run(debug=True)