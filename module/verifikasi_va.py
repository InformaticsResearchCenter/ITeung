from module import va_parent
from module import kelas
from module import surat_va

from datetime import datetime

from openpyxl import load_workbook

import config

def getDataPayment(npm):
    db = va_parent.dbConnectVA()
    data_upload=getTrxIdFromUpload(npm)
    trx_id=data_upload['trx_id']
    sql = f"select * from payment_notification where trx_id = '{trx_id}'"
    with db:
        cur = db.cursor()
        cur.execute(sql)
        data=cur.fetchone()
        if data != None:
            fields = map(lambda x: x[0], cur.description)
            result = dict(zip(fields, data))
        else:
            result=None
        return result

def getTrxIdFromUpload(npm):
    db = va_parent.dbConnectVA()
    sql = f"select * from upload where trx_id like '%INV-%' and trx_id like '%-SPP-%' and trx_id like '%-{npm}-%' and expired_date > CURRENT_DATE"
    with db:
        cur = db.cursor()
        cur.execute(sql)
        data=cur.fetchone()
        if data != None:
            fields = map(lambda x: x[0], cur.description)
            result = dict(zip(fields, data))
        else:
            result=None
        return result

def auth(data):
    if kelas.getNpmandNameMahasiswa(data[0]):
        return True
    else:
        return False

def replymsg(driver, data):
    npm, nama=kelas.getNpmandNameMahasiswa(data[0])
    datava=getDataPayment(npm)
    if datava:
        datenow = datetime.date(datetime.now()).strftime('%d-%m-%Y')
        yearnow = datetime.date(datetime.now()).strftime('%Y')
        trxid = datava['trx_id']
        npm = cekNpmInTrxID(trxid)
        tipesemester = cekTipeSemester(trxid)
        tahunid = f'{yearnow}{tipesemester}'
        prodiid = f'{npm[0]}{npm[3]}'
        virtual_account = datava['virtual_account']
        customer_name =datava['customer_name']
        trx_amount = datava['trx_amount']
        payment_amount = datava['payment_amount']
        cumulative_payment_amount = datava['cumulative_payment_amount']
        payment_ntb = datava['payment_ntb']
        datetime_payment = datava['datetime_payment']
        datetime_payment_iso8601 = datava['datetime_payment_iso8601']
        ws = openfile().active
        prodi_singkatan = kelas.getProdiSingkatanFromProdiID(kelas.getProdiIDwithStudentID(npm)).lower()
        tingkat = f"tk{int(datetime.now().strftime('%Y')) - int(kelas.getTahunAngkatanWithStudentID(npm)) + 1}"
        angkatan = kelas.getTahunAngkatanWithStudentID(npm)
        key = f'{prodi_singkatan}{tingkat}{angkatan}'
        default_amount_payment = getDataDefault(key, ws)
        if default_amount_payment:
            default_amount_payment=int(default_amount_payment)
        else:
            default_amount_payment=2500000
        message = f'Hai haiiiii kamu sudah transfer pembayaran semester yaaaa dengan\n\n*NPM: {npm}*\n*Nama: {customer_name}*\n*Virtual Account: {virtual_account}*\n*Tanggal: {datetime_payment}*\n*Jumlah Transfer: {floatToRupiah(payment_amount)}*\n*Total Sudah Bayar: {floatToRupiah(cumulative_payment_amount)}*\n*Total Harus Bayar: {floatToRupiah(trx_amount)}*\n*Sisa Yang Harus Dibayar: {floatToRupiah(float(int(trx_amount)-int(cumulative_payment_amount)))}*'
        if int(trx_amount) > int(default_amount_payment):
            amount_tunggakan = int(trx_amount) - int(default_amount_payment)
            fifty_percent_default_payment = int(default_amount_payment) / 2
            minimum_payment = int(amount_tunggakan) + int(fifty_percent_default_payment)
        else:
            potongan = int(default_amount_payment) - int(trx_amount)
            minimum_payment = int(default_amount_payment) / 2
            cumulative_payment_amount += potongan
        openfile().close()
        if float(cumulative_payment_amount) >= float(minimum_payment):
            if cekSudahAdaKHS(npm, tahunid, 'A'):
                updateBiayaKHS(npm, tahunid, trx_amount - cumulative_payment_amount)
                message += f'\n\nterima kasih yaaa sudah bayar semester, semangat kuliahnya kakaaaa......'
            else:
                message += f'\n\nKamu *sudah bisa* isi KRS yaaa coba cek di *SIAP* yaaa...., #BOTNAME# ucapkan terima kasihhhh dan jangan salah saat isi KRS yaaa....'
                message = message.replace('#BOTNAME#', config.bot_name)
                insertnewKHS(npm, tahunid, prodiid, cekSesiSemester(tipesemester, npm), trx_amount - cumulative_payment_amount)
        else:
            message += f'\n\nYahhhh kamu *belum bisa* isi KRS nihhhh coba *buat surat* lalu *ajukan ke pihak BAUK* agar kamu bisa isi KRS..... Suratnya udah {config.bot_name} kirim ke *{kelas.getStudentEmail(npm)}*'
            surat_va.makePdfAndSendToEmail(npm)
        msgreply=message
    else:
        msgreply='kamu belum ada transfer'
    return msgreply

def cekNpmInTrxID(trxid):
    data=trxid.split('-')
    for i in data:
        if kelas.getDataMahasiswa(i):
            return i


def cekTipeSemester(trxid):
    for i in trxid.split('-'):
        try:
            int(i)
            isNumber=True
        except:
            isNumber=False
        if len(i) == 1 and isNumber:
            return i


def openfile():
    namafile='wekwek.xlsx'
    wb = load_workbook(namafile)
    return wb


def getDataDefault(key, ws):
    switcher={
        'd3titk22019': ws['F8'].value,
        'd3titk32018': ws['H8'].value,
        'd3mitk22019': ws['F9'].value,
        'd3mitk32018': ws['H9'].value,
        'd3aktk22019': ws['F10'].value,
        'd3aktk32018': ws['H10'].value,
        'd3mbtk22019': ws['F11'].value,
        'd3mbtk32018': ws['H11'].value,
        'd3lbtk22019': ws['F12'].value,
        'd3lbtk32018': ws['H12'].value,
        'd4titk22019': ws['F21'].value,
        'd4titk32018': ws['H21'].value,
        'd4titk42017': ws['J21'].value,
        'd4aktk22019': ws['F22'].value,
        'd4aktk32018': ws['H22'].value,
        'd4aktk42017': ws['J22'].value,
        'd4mbtk22019': ws['F23'].value,
        'd4mbtk32018': ws['H23'].value,
        'd4mbtk42017': ws['J23'].value,
        'd4lbtk22019': ws['F24'].value,
        'd4lbtk32018': ws['H24'].value,
        'd4lbtk42017': ws['J24'].value
    }
    return switcher.get(key, False)

def insertnewKHS(npm, tahunid, prodiid, tipesemester, biaya):
    db=kelas.dbConnectSiap()
    sql=f"INSERT INTO simak_trn_khs (KHSID,TahunID,ProdiID,KodeID,ProgramID,MhswID,StatusMhswID,sesi,MaxSKS,Cetak,Biaya) VALUES (DEFAULT,'{tahunid}','{prodiid} ','YPBPI','REG','{npm}','A','{tipesemester}','24','Y', {biaya});"
    with db:
        cur=db.cursor()
        cur.execute(sql)

def updateBiayaKHS(npm, tahunid, biaya):
    db=kelas.dbConnectSiap()
    sql=f"UPDATE simak_trn_khs SET Biaya={biaya} WHERE MhswID={npm} and TahunID={tahunid}"
    with db:
        cur=db.cursor()
        cur.execute(sql)

def cekSudahAdaKHS(npm, tahunid, statusmahasiswa):
    db=kelas.dbConnectSiap()
    sql=f"select * from simak_trn_khs where MhswID={npm} and TahunID={tahunid} and StatusMhswID='{statusmahasiswa}'"
    with db:
        cur=db.cursor()
        cur.execute(sql)
        row=cur.fetchone()
        if row is not None:
            return True
        else:
            return False

def cekSesiSemester(tipesemester, npm):
    tahunangkatan=int(kelas.getTahunAngkatanWithStudentID(npm))
    tahunsekarang=int(datetime.now().strftime('%Y'))
    selisihtahun=tahunsekarang-tahunangkatan
    if tipesemester == '1':
        selisihtahun+=0.5
        sesisemester=selisihtahun/0.5
    else:
        selisihtahun+=1
        sesisemester=selisihtahun/0.5
    return str(sesisemester)

def floatToRupiah(uang):
    uang=float(uang)
    str_uang=str(uang)
    uang_split=str_uang.split('.')[0]
    uang_reverse=uang_split[::-1]
    uang_temp=''
    for i, j in enumerate(uang_reverse):
        if i%3==0:
            uang_temp+=f'.{j}'
        else:
            uang_temp+=f'{j}'
    uang_reverse_jadi=uang_temp[1:]
    return f'Rp {uang_reverse_jadi[::-1]},{str_uang.split(".")[1]}'
def keyword():
    return ['verifikasi va pembayaran']
