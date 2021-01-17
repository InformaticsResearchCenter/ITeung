from module import kelas
from lib import wa, reply
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
import os
import subprocess
import string
import ssl
import smtplib
import config
import time
import unicodedata
import re
import urllib.request
import sys
import pandas as pd
import pymysql
import threading
import shutil

def auth(data):
    if kelas.getKodeDosen(data[0]) == '':
        ret = False
    else:
        ret = True
    return ret

def replymsg(driver, data):
    if kelas.cekSiap():
        kodedosen = kelas.getKodeDosen(data[0])
        wmsg = reply.getWaitingMessage(os.path.basename(__file__).split('.')[0])
        wmsg = wmsg.replace('#EMAIL#', kelas.getEmailDosen(kodedosen))
        wmsg = wmsg.replace('#BOTNAME#', config.bot_name)
        wa.typeAndSendMessage(driver, wmsg)
        msg = data[3].lower()
        
        try:
            data = f"{kelas.getEmailDosen(kodedosen)};{[npm for npm in msg.split(' ') if npm.isdigit() and len(npm) == 7][0]}"
            subprocess.Popen(["python", "run.py", os.path.basename(__file__).split('.')[0],data], cwd=config.cwd)
        except Exception as e: 
            wa.typeAndSendMessage(driver, 'Mimpam zuzuzu.. anda salah keyword...'+str(e))
        
    else:
        wa.typeAndSendMessage(
            driver, 'Mohon maaf server Akademik SIAP sedang dalam kondisi DOWN, mohon untuk menginformasikan ke ADMIN dan tunggu hingga beberapa menit kemudian, lalu ulangi kembali, terima kasih....')
    return ''

def sendEmail(email, file, mhs):
    try:
        subject = f"Template Daftar Penilaian Sidang TA {mhs}"
        body = f"Jadi ini template daftar penilaian sidang TA {mhs}. Mohon dicek kembali, apabila ada kesalahan pada template ini.\nForm penilaian yg sudah diinputkan, dikirimkan segera hari ini juga ke email koordinator TA woroisti@poltekpos.ac.id atau wistirahayu@gmail.com"

        sender_email = config.email_iteung
        receiver_email = email
        # receiver_email = 'divakrishnam@yahoo.com'
        password = config.pass_iteung

        message = MIMEMultipart()
        message["From"] = f'ITeung <{config.email_iteung}>'
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = receiver_email

        message.attach(MIMEText(body, "plain"))

        with open(f'sidang\\{file}', "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            "attachment; filename= %s " % file,
        )

        message.attach(part)
        
        text = message.as_string()

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)

        print(f'File {file} berhasil dikirim ke {email}')
    except FileNotFoundError:
        print("File tidak ditemukan")
    except Exception as e: 
        print(str(e))
        
def run(data):
    npm = int(data.split(";")[1])
    email = data.split(";")[0]
    
    df = pd.read_excel(f'jadwal_sidang_ta_14.xlsx')
    df.set_index('npm', inplace=True)
    
    listPem = ['pem1', 'pem2', 'pem3', 'pem4', 'jadwal']
    pem = df.loc[npm, listPem]
    pemUt = pem[0]
    pemPen = pem[1]
    penUt = pem[2]
    penPen = pem[3]
    jadwal = pem[4]
    
    namaPemUt = getDataDosen(pemUt)
    namaPemPen = getDataDosen(pemPen)
    namaPenUt = getDataDosen(penUt)
    namaPenPen = getDataDosen(penPen)
    dataMhs = getDataMahasiswa(npm)
    namaMhs = dataMhs[0]
    prodi = dataMhs[1]
    # tahun = '20192'
    tahun = kelas.getTahunID()
    judul = getJudul(npm, tahun)
    
    namaFile = f"bap-sidang-{npm}.xlsx"
    # shutil.copy("sidang\\TEMPLATE DAFTAR PENILAIAN SIDANG TUGAS AKHIR.xlsx", f"sidang\\{namaFile}")
    shutil.copy("sidang\\TEMPLATE DAFTAR PENILAIAN SIDANG INTERNSHIP I.xlsx", f"sidang\\{namaFile}")
    
    wb = load_workbook(filename=f"sidang\\{namaFile}")
    
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Berita Acara':
            wb.active = s
            sheet = wb.active
            sheet["C12"] = namaMhs
            sheet["C13"] = str(npm)
            sheet["C14"] = judul
            sheet["C15"] = namaPemUt
            sheet["C16"] = namaPemPen
            sheet["C17"] = namaPenUt
            sheet["C18"] = namaPenPen
            
            sheet["A27"] = f'( {namaPenUt} )'
            sheet["D27"] = f'( {namaPenPen} )'
            
        elif wb.sheetnames[s] == 'Risalah Penguji Utama':
            wb.active = s
            sheet = wb.active
            sheet["D10"] = namaMhs
            sheet["D11"] = str(npm)
            sheet["D12"] = prodi
            sheet["D13"] = tahun[:4]+"/"+str(int(tahun[:4])+1)
            sheet["D14"] = jadwal
            sheet["D15"] = judul
            # sheet["I81"] = f'( {namaPenUt} )'
            
        elif wb.sheetnames[s] == 'Risalah Penguji Pendamping':
            wb.active = s
            sheet = wb.active
            sheet["D10"] = namaMhs
            sheet["D11"] = str(npm)
            sheet["D12"] = prodi
            sheet["D13"] = tahun[:4]+"/"+str(int(tahun[:4])+1)
            sheet["D14"] = jadwal
            sheet["D15"] = judul
            # sheet["I81"] = f'( {namaPenPen} )'
            
        elif wb.sheetnames[s] == 'Penilaian Penguji Utama':
            wb.active = s
            sheet = wb.active
            sheet["C9"] = namaMhs
            sheet["C10"] = str(npm)
            sheet["C11"] = judul
            # sheet["D32"] = f'( {namaPenUt} )'
            
        elif wb.sheetnames[s] == 'Penilaian Penguji Pendamping':
            wb.active = s
            sheet = wb.active
            sheet["C9"] = namaMhs
            sheet["C10"] = str(npm)
            sheet["C11"] = judul
            # sheet["D32"] = f'( {namaPenPen} )'
            
        else:
            pass
    wb.save(filename=f"sidang\\{namaFile}")
    
    sendEmail(email, namaFile, namaMhs)
    
def getDataDosen(dosenid):
    db= kelas.dbConnectSiap()
    sql=f'select concat(Nama, " ", Gelar) as nama from simak_mst_dosen where Login="{dosenid}"'
    with db:
        cur=db.cursor()
        cur.execute(sql)
        row=cur.fetchone()
        if row is not None:
            return row[0]
        else:
            return None

def getDataMahasiswa(npm):
    db = kelas.dbConnectSiap()
    sql = f"select smm.Nama, smp.Nama from simak_mst_mahasiswa as smm inner join simak_mst_prodi as smp on smm.ProdiID = smp.ProdiID where MhswID = '{npm}'"
    with db:
        cur = db.cursor()
        cur.execute(sql)
        row = cur.fetchone()
        if row is not None:
            return row
        else:
            return None

def getJudul(npm, tahunID):
    db = kelas.dbConnect()
    sql = f"select judul from bimbingan_data where npm ='{npm}' and tahun_id='{tahunID}'"
    with db:
        cur = db.cursor()
        cur.execute(sql)
        row = cur.fetchone()
        if row is not None:
            return row[0]
        else:
            return None