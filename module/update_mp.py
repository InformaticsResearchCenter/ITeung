from module import kelas, cek_mp, hakiaptimas
from lib import wa, reply, message, numbers
from openpyxl import load_workbook
import os, config, pandas

def auth(data):
    if kelas.getKodeDosen(data[0]) == '':
        ret = False
    else:
        ret = True
    return ret

def replymsg(driver, data):
    wmsg = reply.getWaitingMessage(os.path.basename(__file__).split('.')[0])
    wa.typeAndSendMessage(driver, wmsg)
    num=numbers.normalize(data[0])
    msg=message.normalize(data[3])
    msgs=msg.split(' ')
    excelkeyword=msgs[-1]
    dosenID=kelas.getKodeDosen(num)
    if excelkeyword == 'excel':
        try:
            namafile=hakiaptimas.downloadFile(driver)
            downloadstatus=True
        except Exception as e:
            downloadstatus=False
        if downloadstatus:
            if namafile.split('.')[1] == 'xlsx' or namafile.split('.')[1] == 'xls':
                hakiaptimas.moveFiles(namafile)
                msgreply='okeee sudah #BOTNAME# update yaa materi perkuliahannya:'
                wb = load_workbook(namafile)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'jadwal id':
                        continue
                    else:
                        jadwalid = row[0]
                        pertemuan = row[1]
                        materiperkuliahan = row[2]
                        updateMateriPerkuliahan(jadwalid, pertemuan, materiperkuliahan)
                        msgreply+=f'\n\nJadwal ID: {jadwalid}\nPertemuan: {pertemuan}\nMateri Perkuliahan: {materiperkuliahan}'
                deleteFilesWithCWD(namafile)
            else:
                msgreply='format file salah'
                deleteFilesOnDownloadsFolder(namafile)
        else:
            msgreply = 'mana filenya coyyyyy'
    else:
        try:
            datasplit=msg.split(' materi perkuliahan ')[1]
            jadwalid=int(datasplit.split(' ')[0])
            jadwalidDosen=cek_mp.getJadwalIDfromDosenID(dosenID)
            for i in jadwalidDosen:
                if jadwalid in i:
                    isAdaJadwal=True
                    break
                else:
                    isAdaJadwal=False
            if isAdaJadwal:
                pertemuan=datasplit.split(' ')[2]
                if len(pertemuan) > 1:
                    materi=datasplit.split(' pertemuan ')[1][3:]
                else:
                    materi=datasplit.split(' pertemuan ')[1][2:]
                print(f'{jadwalid} | {pertemuan} | {materi}')
                updateMateriPerkuliahan(jadwalid, pertemuan, materi)
                msgreply=f'okeee bosqqq sudah di update yaaa\nJadwal ID: {jadwalid}\nPertemuan: {pertemuan}\nMateri: {materi}'
            else:
                msgreply='Jadwal ID yang dimasukkan salah atau tidak ditemukan!'
        except Exception as e:
            print(str(e))
            msgreply='format katanya salah nichhhh bosqueeee'
    return msgreply

def updateMateriPerkuliahan(jadwalid, pertemuan, materi):
    db=kelas.dbConnectSiap()
    sql=f"UPDATE `simpati`.`simak_trn_presensi_dosen` SET `MP` = '{materi}' WHERE `JadwalID` = {jadwalid} AND `Pertemuan` = {pertemuan} AND `TahunID` = {kelas.getTahunID()}"
    with db:
        cur=db.cursor()
        cur.execute(sql)


def deleteFilesOnDownloadsFolder(namafile):
    source = 'C:\\Users\\' + config.computeruser + '\\Downloads\\' + str(namafile)
    deletefiles = True
    while deletefiles:
        try:
            os.remove(source)
            deletefiles = False
        except:
            deletefiles = True

def deleteFilesWithCWD(namafile):
    source = os.getcwd()+'\\'+str(namafile)
    deletefiles = True
    while deletefiles:
        try:
            os.remove(source)
            deletefiles = False
        except:
            deletefiles = True