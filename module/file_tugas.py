from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from fpdf import FPDF
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys

import os
import subprocess
import string
import ssl
import smtplib
import config
import time
import unicodedata
import re
import urllib.request
import sys
import pandas as pd
import pymysql
import openpyxl
import threading
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from lib import reply
import os
from lib import wa
from module import kelas
from lib import wa, numbers


def auth(data):
    if kelas.getKodeDosen(data[0]) == '':
        ret = False
    else:
        ret = True
    return ret

def replymsg(driver, data):
    if kelas.cekSiap():
        kodedosen = kelas.getKodeDosen(data[0])
        wmsg = reply.getWaitingMessage(os.path.basename(__file__).split('.')[0])
        wmsg = wmsg.replace('#EMAIL#', getEmailDosen(kodedosen))
        wmsg = wmsg.replace('#BOTNAME#', config.bot_name)
        wa.typeAndSendMessage(driver, wmsg)
        num = data[0]
        num = numbers.normalize(num)

        subprocess.Popen(["python", "run.py", os.path.basename(__file__).split('.')[0],num],
                        cwd=config.cwd)        
    else:
        wa.typeAndSendMessage(
            driver, 'Mohon maaf server Akademik SIAP sedang dalam kondisi DOWN, mohon untuk menginformasikan ke ADMIN dan tunggu hingga beberapa menit kemudian, lalu ulangi kembali, terima kasih....')
    return ''

def run(num):
    param = {
                'nomor': num,
                'tahun': kelas.getTahunID(),
            }
    makeExcelAndSend(param)
    
# def replymsg(driver, data):
#     if kelas.cekSiap():
#         # wmsg = reply.getWaitingMessage(os.path.basename(__file__).split('.')[0])
#         # wa.typeAndSendMessage(driver, wmsg)
#         num = data[0]
#         num = numbers.normalize(num)
#         msg = data[3]
#         data = msg.split(' ')
#         try:
#             param = {
#                 'nomor': num,
#                 'tahun': kelas.getTahunID(),
#             }
#             makeExcelAndSend(param)
#             msgreply = 'Ditunggu bos.....\nNtar dikirim ke email semua filenya....'
#         except:
#             msgreply = 'Salah keyword beb....., atau salah masukin jadwal uas atau uts'
#     else:
#         msgreply = 'Mohon maaf server Akademik SIAP sedang dalam kondisi DOWN, mohon untuk menginformasikan ke ADMIN dan tunggu hingga beberapa menit kemudian, lalu ulangi kembali, terima kasih....'
#     return msgreply


def dbConnectSiap():
    db = pymysql.connect(config.db_host_siap,
                         config.db_username_siap,
                         config.db_password_siap,
                         config.db_name_siap)
    return db


def checkDosen(nomor, tahun):
    db = dbConnectSiap()
    query = """
            select distinct(Nama) from simak_trn_jadwal where DosenID = (select Login from simak_mst_dosen where Handphone = '""" + nomor + """') and TahunID = '""" + tahun + """'
        """
    with db:
        cur = db.cursor()
        cur.execute(query)
        rows = cur.fetchall()
        if cur.rowcount > 0:
            for row in rows:
                return row[0]
        else:
            return False


def getHeaderTugas(jadwalID):
    db = dbConnectSiap()
    sql = """
            select j.JadwalID, j.MKKode, j.Nama, j.NamaKelas, concat(d.Nama, ', ', d.Gelar) as Pengajar, time_format(j.JamMulai, '%H:%i') as JamMulai, time_format(j.JamSelesai, '%H:%i') as JamSelesai, r.Nama as Ruang, j.JumlahMhsw from simak_trn_jadwal j, simak_mst_dosen d, simak_mst_matakuliah m, simak_mst_ruangan r, simak_mst_tahun t, simak_mst_prodi pr where j.JadwalID = '"""+jadwalID+"""' and j.MKID = m.MKID and j.DosenID = d.Login and j.RuangID = r.RuangID and j.TahunID = t.TahunID and t.ProgramID = 'REG' and m.ProdiID = pr.ProdiID group by j.JadwalID;
        """

    with db:
        cur = db.cursor()
        cur.execute(sql)
        row = cur.fetchone()
        if row is not None:
            header = {
                'kode_matkul': row[1]+' / '+row[2],
                'kelas': convertKelas(int(row[3].strip("0"))),
                'pengajar': row[4],
                'jadwal_id': str(row[0]),
                'jadwal_ruang': row[5]+' - '+row[6]+' / '+row[7],
                'peserta': str(row[8])
            }
            return header
        else:
            return False


def getBodyTugas(jadwalID):
    db = dbConnectSiap()
    query = """
            select krs.MhswID, mhs.Nama, krs.Tugas1, krs.Tugas2, krs.Tugas3, krs.Tugas4, krs.Tugas5 from simak_trn_krs krs, simak_mst_mahasiswa mhs, simak_trn_jadwal j where krs.StatusKRSID='A' and krs.JadwalID='"""+jadwalID+"""' and krs.TahunID=j.TahunID and krs.NA='N' and krs.MhswID=mhs.MhswID group by krs.MhswID order by krs.MhswID ASC
        """
    with db:
        body = []
        cur = db.cursor()
        cur.execute(query)
        rows = cur.fetchall()
        if rows is not None:
            for row in rows:
                body.append([row[0], row[1], int(row[2]), int(
                    row[3]), int(row[4]), int(row[5]), int(row[6])])
            return body
        else:
            return False


def getMatkulTugas(nomor, tahun):
    db = dbConnectSiap()
    sql = """
select j.JadwalID, j.Nama, CASE
        WHEN j.NamaKelas =1 THEN 'A'
        WHEN j.NamaKelas =2 THEN 'B'
        WHEN j.NamaKelas =3 THEN 'C'
        WHEN j.NamaKelas =4 THEN 'D'
        WHEN j.NamaKelas =5 THEN 'E'
        WHEN j.NamaKelas =6 THEN 'F'
        WHEN j.NamaKelas =7 THEN 'G'
        WHEN j.NamaKelas =8 THEN 'H'
        WHEN j.NamaKelas =9 THEN 'I'
        END AS namakelas, d.Email, CASE
        WHEN j.ProdiID ='.13.' THEN 'D3 Teknik Informatika'
        WHEN j.ProdiID ='.14.' THEN 'D4 Teknik Informatika'
        WHEN j.ProdiID ='.23.' THEN 'D3 Manajemen Informatika'
        WHEN j.ProdiID ='.33.' THEN 'D3 Akuntansi'
        WHEN j.ProdiID ='.34.' THEN 'D4 Akuntansi Keuangan'
        WHEN j.ProdiID ='.43.' THEN 'D3 Manajemen Pemasaran'
        WHEN j.ProdiID ='.44.' THEN 'D4 Manajemen Perusahaan'
        WHEN j.ProdiID ='.53.' THEN 'D3 Logistik Bisnis'
        WHEN j.ProdiID ='.54.' THEN 'D4 Logistik Bisnis'
        END AS namaprodi
        from simak_trn_jadwal j, simak_mst_dosen d, simak_mst_matakuliah m,  simak_mst_tahun t, simak_mst_prodi pr
        where j.MKID=m.MKID and j.DosenID=d.Login and j.DosenID = (select Login from simak_mst_dosen where Handphone = '"""+nomor+"""') 
        and j.TahunID='"""+tahun+"""' and t.ProgramID='REG' and m.ProdiID = pr.ProdiID group by j.JadwalID;
    """
    with db:
        jadwal = []

        cur = db.cursor()
        cur.execute(sql)
        rows = cur.fetchall()

        if rows is not None:
            for row in rows:
                jadwal.append(list(row))
            return pd.DataFrame(jadwal, columns=['jadwal_id', 'matkul', 'kelas', 'email', 'prodi'])
        else:
            return False


def makeExcelAndSend(param):
    jadwal = getMatkulTugas(
        param['nomor'], param['tahun'])

    for prodi in jadwal['prodi'].unique():
        checkDir(prodi)

    send = list()
    for i in range(len(jadwal)):
        nama_file = 'Penilaian-%s-%s-%s' % (
            changeSpecialChar(jadwal.loc[i, 'prodi']),
            changeSpecialChar(jadwal.loc[i, 'matkul']),
            jadwal.loc[i, 'kelas'])

        data = {
            'jadwal_id': str(jadwal.loc[i, 'jadwal_id']),
            'prodi': jadwal.loc[i, 'prodi'],
            'nama_file': nama_file,
            'semester': convertTahun(param['tahun']),
            'matkul': changeSpecialChar(jadwal.loc[i, 'matkul'])
        }

        if not any(d['jadwal_id'] == jadwal.loc[i, 'jadwal_id'] for d in send):
            t = threading.Thread(target=makeExcel, args=(data,))
            send.append({
                'thread': t,
                'matkul': jadwal.loc[i, 'matkul'],
                'kelas': jadwal.loc[i, 'kelas'],
                'prodi': jadwal.loc[i, 'prodi'],
                'tujuan': jadwal.loc[i, 'email'],
                'nama_file': nama_file+'.xlsx',
                'jadwal_id': jadwal.loc[i, 'jadwal_id']
            })
            t.start()
        else:
            continue

    for data in send:
        data['thread'].join()
        t = threading.Thread(target=sendEmail, args=(data,))
        t.start()


def makeExcel(param):
    head_data = getHeaderTugas(param['jadwal_id'])
    body_data = getBodyTugas(param['jadwal_id'])

    if head_data and body_data:
        title_data = ['Penilaian Tugas ' +
                      param['matkul'].title().replace("_", " ")+" "+param['prodi']+' Semester '+param['semester']]
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 13
        sheet.column_dimensions['C'].width = 50
        generateHead(title_data, head_data, sheet)
        generateBody(body_data, sheet)
        wb.save(
            'tugas/{}/{}.xlsx'.format(param['prodi'], param['nama_file']))
        print('File %s.xlsx berhasil dibuat' % param['nama_file'])
    else:
        pass


def generateHead(title_data, head_data, sheet):
    index_cell = 2
    for title in title_data:
        sheet.merge_cells('A{}:H{}'.format(index_cell, index_cell))
        value = title
        cell = sheet.cell(row=index_cell, column=1)
        cell.value = value
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=16, bold=True)
        index_cell += 1

    sheet.merge_cells('A5:B5')
    cell = sheet.cell(row=5, column=1)
    cell.value = 'Kode / Mata Kuliah'
    cell.font = Font(bold=True)
    sheet["C5"].value = ': '+head_data['kode_matkul']

    sheet.merge_cells('D5:E5')
    cell = sheet.cell(row=5, column=4)
    cell.value = 'Jadwal ID'
    cell.font = Font(bold=True)
    sheet.merge_cells('F5:H5')
    cell = sheet.cell(row=5, column=6)
    cell.value = ': '+head_data['jadwal_id']

    sheet.merge_cells('A6:B6')
    cell = sheet.cell(row=6, column=1)
    cell.value = 'Kelas'
    cell.font = Font(bold=True)
    sheet["C6"].value = ': '+head_data['kelas']

    sheet.merge_cells('D6:E6')
    cell = sheet.cell(row=6, column=4)
    cell.value = 'Jadwal / Ruang'
    cell.font = Font(bold=True)
    sheet.merge_cells('F6:H6')
    cell = sheet.cell(row=6, column=6)
    cell.value = ': '+head_data['jadwal_ruang']

    sheet.merge_cells('A7:B7')
    cell = sheet.cell(row=7, column=1)
    cell.value = 'Pengajar'
    cell.font = Font(bold=True)
    sheet["C7"].value = ': '+head_data['pengajar']

    sheet.merge_cells('D7:E7')
    cell = sheet.cell(row=7, column=4)
    cell.value = 'Peserta'
    cell.font = Font(bold=True)
    sheet.merge_cells('F7:H7')
    cell = sheet.cell(row=7, column=6)
    cell.value = ': '+head_data['peserta']


def generateBody(body_data, sheet):
    index_cell = 9
    cell = sheet["A"+str(index_cell)]
    cell.value = 'No.'
    cell.font = Font(bold=True)
    cell = sheet["B"+str(index_cell)]
    cell.value = 'NPM'
    cell.font = Font(bold=True)
    cell = sheet["C"+str(index_cell)]
    cell.value = 'Nama'
    cell.font = Font(bold=True)
    cell = sheet["D"+str(index_cell)]
    cell.value = 'Tugas 1'
    cell.font = Font(bold=True)
    cell = sheet["E"+str(index_cell)]
    cell.value = 'Tugas 2'
    cell.font = Font(bold=True)
    cell = sheet["F"+str(index_cell)]
    cell.value = 'Tugas 3'
    cell.font = Font(bold=True)
    cell = sheet["G"+str(index_cell)]
    cell.value = 'Tugas 4'
    cell.font = Font(bold=True)
    cell = sheet["H"+str(index_cell)]
    cell.value = 'Tugas 5'
    cell.font = Font(bold=True)
    index_cell += 1

    for body in body_data:
        cell = sheet["A"+str(index_cell)]
        cell.value = index_cell-9
        cell.alignment = Alignment(horizontal='right')
        sheet["B"+str(index_cell)].value = body[0]
        sheet["C"+str(index_cell)].value = body[1]
        cell = sheet["D"+str(index_cell)]
        cell.value = body[2]
        cell.alignment = Alignment(horizontal='right')
        cell = sheet["E"+str(index_cell)]
        cell.value = body[3]
        cell.alignment = Alignment(horizontal='right')
        cell = sheet["F"+str(index_cell)]
        cell.value = body[4]
        cell.alignment = Alignment(horizontal='right')
        cell = sheet["G"+str(index_cell)]
        cell.value = body[5]
        cell.alignment = Alignment(horizontal='right')
        cell = sheet["H"+str(index_cell)]
        cell.value = body[6]
        cell.alignment = Alignment(horizontal='right')
        index_cell += 1


def sendEmail(file):
    try:
        subject = "Penilaian Tugas Mata Kuliah {} Kelas {} Prodi {}".format(
            file['matkul'], file['kelas'], file['prodi'])
        body = "Ini file penilaian oleh iteung ya..., mohon untuk dicek kembali filenya jika ada yang salah mohon untuk diinformasikan ke admin iteung yaa....:) \Penilaian Tugas Mata Kuliah {} Kelas {} Prodi {}".format(
            file['matkul'], file['kelas'], file['prodi'])

        sender_email = config.email_iteung
        receiver_email = file['tujuan']
        # print(receiver_email)
        # receiver_email = 'divakrishnam@yahoo.com'
        password = config.pass_iteung

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = receiver_email

        message.attach(MIMEText(body, "plain"))

        absensifile = file['nama_file']

        with open('tugas\\'+file['prodi']+'\\'+absensifile, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            "attachment; filename= %s " % absensifile,
        )

        message.attach(part)

        text = message.as_string()

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)

        print('File %s berhasil dikirim ke %s' % (absensifile, file['tujuan']))

    except FileNotFoundError:
        pass


def changeSpecialChar(text):
    return text.replace(
        " ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("+", "_").replace("/", "_")


def checkDir(prodi):
    path = 'tugas/{}'.format(prodi)
    if not os.path.exists(path):
        t = threading.Thread(target=os.makedirs, args=(path,))
        t.start()
        t.join()
        print('Direktori {} telah dibuat'.format(prodi))


def convertKelas(kelas):
    list_kelas = list(string.ascii_lowercase)
    list_nomor = list(range(1, 27))
    dict_kelas = dict(zip(list_nomor, list_kelas))
    for k, v in dict_kelas.items():
        if k == kelas:
            return v.upper()
            break
    return 'Kelas tidak terdaftar'


def convertTahun(param):
    tahun = param[:4]
    semester = int(param[-1])
    if semester == 1:
        semester = 'Ganjil'
    elif semester == 2:
        semester = 'Genap'
    return '%s %s/%d' % (semester, tahun, int(tahun)+1)

def getEmailDosen(dosenid):
    db=dbConnectSiap()
    sql="select Email from simak_mst_dosen where Login='{lecturercode}'".format(lecturercode=dosenid)
    with db:
        cur=db.cursor()
        cur.execute(sql)
        rows=cur.fetchone()
        if rows is not None:
            return rows[0]
        else:
            return ''